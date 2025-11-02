from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from zipfile import ZipFile, ZIP_DEFLATED
from datetime import datetime

MAX_PHOTOS = 2  # export up to the first 2 photos per observation


def _flatten_photos_for_export(row):
    """Return dict with photo1_* and photo2_* keys (may be empty)."""
    out = {}
    photos = row.get("photos", []) or []
    for i in range(1, MAX_PHOTOS + 1):
        p = photos[i - 1] if i - 1 < len(photos) else {}
        out[f"photo{i}_name"] = p.get("photoname", "")
        out[f"photo{i}_lat"]  = p.get("photolat", "")
        out[f"photo{i}_lon"]  = p.get("photolon", "")
        out[f"photo{i}_acc"]  = p.get("photoacc", "")
        out[f"photo{i}_dir"]  = p.get("photodir", "")
    return out


def to_excel(rows_dicts, out_path):
    """Simple, single-sheet export of whatever dicts you pass in."""
    df = pd.DataFrame(rows_dicts)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Observations")
    return str(out)


def to_excel_multisheet(observation_rows, out_path):
    """Multi-sheet: Observations + Photos (long form)."""
    obs_clean = []
    photos_rows = []
    for r in observation_rows:
        rc = dict(r)
        photos = rc.pop("photos", [])
        obs_clean.append(rc)
        for p in photos:
            pr = {
                "source_file": rc.get("source_file", ""),
                "seqno": rc.get("seqno", ""),
                "observer": rc.get("observer", ""),
                "event_timestamp": rc.get("event_timestamp", ""),
                "photoname": p.get("photoname", ""),
                "photolat": p.get("photolat", ""),
                "photolon": p.get("photolon", ""),
                "photoacc": p.get("photoacc", ""),
                "photodir": p.get("photodir", ""),
            }
            photos_rows.append(pr)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        pd.DataFrame(obs_clean).to_excel(xw, index=False, sheet_name="Observations")
        if photos_rows:
            pd.DataFrame(photos_rows).to_excel(xw, index=False, sheet_name="Photos")
    return str(out)


def to_excel_with_photo_dropdown(observation_rows, selected_indices, out_path):
    """
    Observations sheet includes:
      - all normal columns
      - photo1_* and photo2_* columns
      - PhotoChoice (Excel dropdown)
      - PhotoValue (formula based on PhotoChoice)
    `selected_indices` is a list of row indices from observation_rows to export (or None for all).
    """
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    # 1) Build rows
    export_rows = []
    indices = selected_indices if selected_indices else range(len(observation_rows))
    for idx in indices:
        base = dict(observation_rows[idx])
        base.pop("photos", None)  # keep long sheet for full photo info
        flat = _flatten_photos_for_export(observation_rows[idx])
        base.update(flat)
        export_rows.append(base)

    if not export_rows:
        wb = Workbook()
        ws = wb.active
        ws.title = "Observations"
        wb.save(out)
        return str(out)

    # 2) Write with pandas first
    df = pd.DataFrame(export_rows)
    photo_cols = [c for c in df.columns if c.startswith("photo1_") or c.startswith("photo2_")]
    base_cols = [c for c in df.columns if c not in photo_cols]
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        df[base_cols + photo_cols].to_excel(xw, index=False, sheet_name="Observations")
        # Long photos sheet
        photos_long = []
        for idx in indices:
            r = observation_rows[idx]
            for i, p in enumerate(r.get("photos", []), start=1):
                photos_long.append({
                    "source_file": r.get("source_file", ""),
                    "seqno": r.get("seqno", ""),
                    "photo_index": i,
                    "photoname": p.get("photoname", ""),
                    "photolat": p.get("photolat", ""),
                    "photolon": p.get("photolon", ""),
                    "photoacc": p.get("photoacc", ""),
                    "photodir": p.get("photodir", ""),
                })
        if photos_long:
            pd.DataFrame(photos_long).to_excel(xw, index=False, sheet_name="Photos")

    # 3) Post-process with openpyxl: add dropdown + formula
    wb = load_workbook(out)
    ws = wb["Observations"]

    max_col = ws.max_column
    choice_col = max_col + 1
    value_col = max_col + 2
    ws.cell(row=1, column=choice_col, value="PhotoChoice")
    ws.cell(row=1, column=value_col, value="PhotoValue")

    # Hidden list sheet
    opts = [
        "photo1 name", "photo1 lat", "photo1 lon",
        "photo2 name", "photo2 lat", "photo2 lon",
    ]
    if "Lists" in wb.sheetnames:
        lists_ws = wb["Lists"]
    else:
        lists_ws = wb.create_sheet("Lists")
        lists_ws.sheet_state = "hidden"
    for i, opt in enumerate(opts, start=1):
        lists_ws.cell(row=i, column=1, value=opt)
    dv = DataValidation(type="list", formula1="=Lists!$A$1:$A$6", allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv)

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    c_p1n = headers.get("photo1_name")
    c_p1la = headers.get("photo1_lat")
    c_p1lo = headers.get("photo1_lon")
    c_p2n = headers.get("photo2_name")
    c_p2la = headers.get("photo2_lat")
    c_p2lo = headers.get("photo2_lon")

    col_choice_letter = get_column_letter(choice_col)

    for r in range(2, ws.max_row + 1):
        dv.add(ws.cell(row=r, column=choice_col))
        def _cell(col_idx):
            from openpyxl.utils import get_column_letter as _g
            return f"{_g(col_idx)}{r}" if col_idx else '""'
        formula = (
            f'=IF({col_choice_letter}{r}="photo1 name",{_cell(c_p1n)},'
            f'IF({col_choice_letter}{r}="photo1 lat",{_cell(c_p1la)},'
            f'IF({col_choice_letter}{r}="photo1 lon",{_cell(c_p1lo)},'
            f'IF({col_choice_letter}{r}="photo2 name",{_cell(c_p2n)},'
            f'IF({col_choice_letter}{r}="photo2 lat",{_cell(c_p2la)},'
            f'IF({col_choice_letter}{r}="photo2 lon",{_cell(c_p2lo)},""))))))'
        )
        ws.cell(row=r, column=value_col, value=formula)

    ws.column_dimensions[get_column_letter(choice_col)].width = 18
    ws.column_dimensions[get_column_letter(value_col)].width = 28
    wb.save(out)
    return str(out)


# ------------------ Spatial: single-point KML/KMZ ------------------ #

def _kml_escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def kml_for_point(lat, lon, name="", description=""):
    name = _kml_escape(name)
    description = _kml_escape(description)
    # KML expects lon,lat order in <coordinates>
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
  <name>{name}</name>
  <Placemark>
    <name>{name}</name>
    <description>{description}</description>
    <Point>
      <coordinates>{lon},{lat},0</coordinates>
    </Point>
  </Placemark>
</Document>
</kml>
"""


def save_point_kml(lat, lon, name, description, out_path):
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    xml = kml_for_point(lat, lon, name=name, description=description)
    out.write_text(xml, encoding="utf-8")
    return str(out)


def save_point_kmz(lat, lon, name, description, out_path):
    """Write a KMZ (ZIP) containing doc.kml."""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    xml = kml_for_point(lat, lon, name=name, description=description)
    with ZipFile(out, "w", compression=ZIP_DEFLATED) as zf:
        zf.writestr("doc.kml", xml.encode("utf-8"))
    return str(out)
